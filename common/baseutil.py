#!/usr/bin/python3
# -*- coding: utf-8 -*-
# @Author: Samuel

import os
import time
import configparser
import yaml
import openpyxl
import csv
import random
import string

root_path = os.path.split(os.path.abspath(os.path.dirname(__file__)))[0]
cur_date = time.strftime('%Y-%m-%d')
cur_time = time.strftime('%H:%M:%S')
cur_datetime = time.strftime('%Y-%m-%d %H:%M:%S')


def get_config_value(section, key):
    """获取config文件中section的key值"""
    config_path = os.path.join(root_path, 'conf', 'config.ini')
    config = configparser.ConfigParser()
    config.read(config_path, encoding='utf-8')
    value = config.get(section, key)
    return value


def get_yaml_value(option, key):
    """获取yaml配置文件中option的key值"""
    yaml_path = os.path.join(root_path, 'conf', 'browser.yaml')
    with open(yaml_path, 'r', encoding='utf-8') as f:
        temp = yaml.load(f.read(), Loader=yaml.FullLoader)
    value = temp[option][key]
    return value


def get_excel_data(excel_name, sheet_name, column_num):
    """返回excel中第column_num列的值"""
    file_path = os.path.join(root_path, 'files', excel_name)
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    array = []
    # 从Excel第二行开始，第一行为列名
    for i in range(2, sheet.max_row + 1):
        date = sheet.cell(row=i, column=int(column_num)).value
        if date is None:
            break
        else:
            array.append(date)
    return array


def get_every_row_data(excel_name, sheet_name):
    """获取Excel中每一行的值，返回一个二维数组"""
    file_path = os.path.join(root_path, 'files', excel_name)
    wb = openpyxl.load_workbook(file_path)
    sheet = wb[sheet_name]
    array = []
    for r in range(1, sheet.max_row + 1):
        date = (''.join([str(sheet.cell(row=r, column=c).value).ljust(10) for c in range(1, sheet.max_column + 1)]))
        if date.split()[0] == 'None':
            break
        else:
            array.append(date.split())
    return array


def get_csv_data(csv_name):
    """返回一个二维数组"""
    csv_path = os.path.join(root_path, 'files', csv_name)
    with open(csv_path) as csv_file:
        read_csv = csv.reader(csv_file, delimiter=',')
        array = []
        for row in read_csv:
            if len(row):
                array.append(row)
            else:
                break
    return array


def verify_in(small, big):
    """判断一个数据类型是否存在于另一个类型中"""
    if isinstance(small, dict):
        for key in small:
            if key in big:
                if isinstance(small[key], dict):
                    if verify_in(small[key], big[key]):
                        continue
                    else:
                        return False
                elif isinstance(small[key], list):
                    for sml_list, big_list in zip(small[key], big[key]):
                        if verify_in(sml_list, big_list):
                            continue
                        else:
                            return False
                else:
                    if str(small[key]) in str(big[key]):
                        continue
                    else:
                        return False
            else:
                return False
        return True
    if isinstance(small, list):
        for i in small:
            if i in big:
                continue
            else:
                return False
        return True
    else:
        if str(small) in str(big):
            return True
        else:
            return False


def get_folders(file_path):
    """获取file_path的所有目录名"""
    ff_list = os.listdir(file_path)
    for item in ff_list[::]:
        if os.path.isfile(os.path.join(file_path, item)):
            ff_list.remove(item)
    return ff_list


def get_dict_values(search_dict, search_key):
    """
    查找一个字典的key的所有键值
    """
    found_values = []
    for key, value in search_dict.items():
        if key == search_key:
            found_values.append(value)
        elif isinstance(value, dict):
            results = get_dict_values(value, search_key)
            for result in results:
                found_values.append(result)
        elif isinstance(value, list):
            for item in value:
                if isinstance(item, dict):
                    more_results = get_dict_values(item, search_key)
                    for another_result in more_results:
                        found_values.append(another_result)
    return found_values


def gen_random_string(str_len=10):
    return ''.join(random.sample(string.ascii_letters + string.digits, str_len))
